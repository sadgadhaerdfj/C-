import openpyxl


class ReadFiles(object):
    def __init__(self,workbook,worksheet,txt):
        self.workbook=workbook
        self.worksheet=worksheet
        self.txt=txt
        self.wc=openpyxl.load_workbook(workbook+'.xlsx')
        self.ws=self.wc[self.worksheet]
        self.row=self.ws.max_row
        self.column=self.ws.max_column
        self.f=open(self.txt+'.txt','w')
        self.write_in()

    def write_in(self):
        for i in range(self.row):
            for j in range(self.column):
                cell_value=self.ws.cell(row=i+1,column=j+1).value
                if(cell_value is not None):
                    self.f.write(str(cell_value)+"\t")
                else:
                    self.f.write("\t")
            self.f.write("\n")
        self.f.close()

read=ReadFiles("（前3次）2022级导论考试--成绩表","598854","598854")
read1=ReadFiles("（前3次）2022级导论考试--成绩表","598856","598856")
read2=ReadFiles("（前3次）2022级导论考试--成绩表","598858","598858")